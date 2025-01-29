import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

class ExcelWriter:
    def __init__(self, file_path=os.environ.get("RESULTS_EXCEL_PATH"), config=None):
        self.file_path = file_path
        self.config = config or {}
        self._ensure_directory_exists()

    def _ensure_directory_exists(self):
        directory = os.path.dirname(self.file_path)
        if not os.path.exists(directory):
            os.makedirs(directory)

    def write_dataframe(self, df, sheet_name='Default_Evaluation'):
        version_number = datetime.now().strftime("%Y-%m-%dT%H:%M")
        question_source = os.environ.get('DATASET_FILENAME', 'unknown')
        doc_format = os.path.splitext(question_source)[1].replace(".", "")
        number_of_questions = len(df)
        embedding_model = self.config.get('embedding', {}).get('model', 'unknown')
        evaluated_model = self.config.get('llm_to_be_evaluated', {}).get('model', 'unknown')
        judge_model = self.config.get('llm_judge', {}).get('model', 'unknown')

        additional_data = {
            'Version number': [version_number] + [''] * (len(df) - 1),
            'Question source': [question_source] + [''] * (len(df) - 1),
            'Doc format': [doc_format] + [''] * (len(df) - 1),
            'Number of questions': [number_of_questions] + [''] * (len(df) - 1),
            'Embedding model': [embedding_model] + [''] * (len(df) - 1),
            'Evaluated model': [evaluated_model] + [''] * (len(df) - 1),
            'Judge model': [judge_model] + [''] * (len(df) - 1)
        }

        for key, value in additional_data.items():
            df[key] = value

        if 'score' in df.columns:
            average_score = df['score'].mean()
            df['average'] = [average_score] + [''] * (len(df) - 1)

        separator = pd.DataFrame([['-' * 10] * len(df.columns)], columns=df.columns)
        df = pd.concat([df, separator], ignore_index=True)

        if not os.path.exists(self.file_path):
            df.to_excel(self.file_path, sheet_name=sheet_name, index=False)
        else:
            with pd.ExcelWriter(self.file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                writer._book = load_workbook(self.file_path)
                if sheet_name in writer.book.sheetnames:
                    startrow = writer.book[sheet_name].max_row
                else:
                    startrow = 0
                df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=startrow==0)
